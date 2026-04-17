"""外注検査 JLAC10 データ取得ツール CLI"""

import argparse
import json
import logging
import sys
from pathlib import Path

import requests

from .categories import CATEGORIES
from .scraper import check_update_needed, diff_report, scrape_all as srl_scrape_all
from .bml import check_bml_update_needed, scrape_all as bml_scrape_all
from .lsi import check_lsi_update_needed, scrape_all as lsi_scrape_all
from .merge import merge_all
from .jslm import scrape_all as jslm_scrape_all, check_jslm_update_needed
from .search import build_index, format_results
from .reagent import build_reagent_db, add_pmda_to_db
from .sop_parser import parse_sop, parse_sop_directory
from .converter import convert_tabular, convert_auto, write_jlac10_to_excel
from .vendor_profiles import list_vendors
from .mapper import bulk_map, export_mapping_excel, export_mapping_json
from .merge import apply_mapping_results
from .ncda_checker import batch_check as ncda_batch_check, export_check_excel


def setup_logging(verbose: bool = False) -> None:
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(
        level=level,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )


# ---------------------------------------------------------------------------
# SRL
# ---------------------------------------------------------------------------

def cmd_srl(args: argparse.Namespace) -> int:
    category_ids = None
    if args.categories:
        category_ids = [int(c) for c in args.categories.split(",")]

    result = srl_scrape_all(
        category_ids=category_ids,
        output_dir=Path(args.output),
        use_cache=not args.no_cache,
        cache_max_age_hours=args.cache_age,
        check_update=args.check_update,
        force=args.force,
    )

    if result is None:
        print("\nSRL: 更新なし。スキップしました。")
        return 0

    meta = result["metadata"]
    print(f"\nSRL 取得完了: {meta['total_items']}件 ({meta['total_categories']}カテゴリ)")
    print(f"  サーバー: {meta['fetched_from_server']}件 / キャッシュ: {meta['served_from_cache']}件")
    if meta.get("last_update_date"):
        print(f"  Last Up Date: {meta['last_update_date']}")
    if meta["errors"]:
        print(f"  エラー: {len(meta['errors'])}件")
        return 1
    return 0


# ---------------------------------------------------------------------------
# BML
# ---------------------------------------------------------------------------

def cmd_bml(args: argparse.Namespace) -> int:
    result = bml_scrape_all(
        output_dir=Path(args.output),
        use_cache=not args.no_cache,
        cache_max_age_hours=args.cache_age,
        check_update=args.check_update,
    )

    if result is None:
        print("\nBML: 更新なし。スキップしました。")
        return 0

    meta = result["metadata"]
    print(f"\nBML 取得完了: {meta['total_items']}件 ({meta['total_categories']}カテゴリ)")
    print(f"  サーバー: {meta['fetched_from_server']}件 / キャッシュ: {meta['served_from_cache']}件")
    if meta["errors"]:
        print(f"  エラー: {len(meta['errors'])}件")
        return 1
    return 0


# ---------------------------------------------------------------------------
# LSI
# ---------------------------------------------------------------------------

def cmd_lsi(args: argparse.Namespace) -> int:
    result = lsi_scrape_all(
        output_dir=Path(args.output),
        use_cache=not args.no_cache,
        check_update=args.check_update,
        cache_max_age_hours=args.cache_age,
    )

    if result is None:
        print("\nLSI: 更新なし。スキップしました。")
        return 0

    meta = result["metadata"]
    print(f"\nLSI 取得完了: {meta['total_items']}件 ({meta['total_list_pages']}リストページ)")
    print(f"  詳細ページ サーバー: {meta['detail_pages_fetched']}件 / キャッシュ: {meta['detail_pages_cached']}件")
    if meta["errors"]:
        print(f"  エラー: {len(meta['errors'])}件")
        return 1
    return 0


# ---------------------------------------------------------------------------
# JSLM
# ---------------------------------------------------------------------------

def cmd_jslm(args: argparse.Namespace) -> int:
    result = jslm_scrape_all(
        output_dir=Path(args.output),
        check_update=args.check_update,
    )

    if result is None:
        print("\nJSLM: 更新なし。スキップしました。")
        return 0

    meta = result["metadata"]
    print(f"\nJSLM JLAC10マスター取得完了 (version: {meta['version']})")
    for k, v in meta["counts"].items():
        print(f"  {k}: {v}件")
    return 0


# ---------------------------------------------------------------------------
# 試薬DB
# ---------------------------------------------------------------------------

def cmd_reagent(args: argparse.Namespace) -> int:
    output_dir = Path(args.output)

    if args.pmda:
        # PMDA URLを追加
        detail = add_pmda_to_db(args.pmda, output_dir)
        print(f"\nPMDA添付文書追加:")
        print(f"  販売名: {detail.get('product_name', '')}")
        print(f"  使用目的: {detail.get('purpose', '')}")
        print(f"  測定原理: {detail.get('principle', '')[:100]}...")
        return 0

    # メーカー製品一覧取得
    result = build_reagent_db(output_dir=output_dir)
    meta = result["metadata"]
    print(f"\n試薬DB構築完了: {meta['total_reagents']}試薬, {meta['total_pmda']}PMDA添付文書")
    return 0


# ---------------------------------------------------------------------------
# SOP パーサー
# ---------------------------------------------------------------------------

def cmd_sop(args: argparse.Namespace) -> int:
    target = Path(args.path)
    output_dir = Path(args.output)

    if target.is_dir():
        result = parse_sop_directory(target, output_dir)
        meta = result["metadata"]
        print(f"\nSOP パース完了: {meta['total_files']}件")
        print(f"  測定法あり: {meta['with_method']}件")
        print(f"  エラー: {meta['errors']}件")
        print(f"  出力: {output_dir}/sop_parsed.json")
    elif target.is_file():
        info = parse_sop(target)
        print(f"\nSOP パース結果: {target.name}")
        print(f"  検査項目: {info['test_item'][:60] or '(未検出)'}")
        print(f"  測定法:   {info['method_summary'][:80] or '(未検出)'}")
        print(f"  試薬:     {info['reagent'][:80] or '(未検出)'}")
        print(f"  装置:     {info['instrument'][:60] or '(未検出)'}")
        print(f"  検体:     {info['specimen'][:60] or '(未検出)'}")
        if info.get("error"):
            print(f"  エラー:   {info['error']}")

        # JSON出力
        out = output_dir / f"sop_{target.stem}.json"
        output_dir.mkdir(parents=True, exist_ok=True)
        out.write_text(json.dumps(info, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"  出力: {out}")
    else:
        print(f"ファイル/ディレクトリが見つかりません: {target}", file=sys.stderr)
        return 1
    return 0


# ---------------------------------------------------------------------------
# Convert (Excel/CSV → JSON)
# ---------------------------------------------------------------------------

def cmd_convert(args: argparse.Namespace) -> int:
    filepath = Path(args.input)

    # column_map を構築
    column_map: dict[str, str] = {}
    if args.col_item:
        column_map["item_name"] = args.col_item
    if args.col_jlac10:
        column_map["jlac10"] = args.col_jlac10
    if args.col_abbr:
        column_map["abbreviation"] = args.col_abbr
    if args.col_std_name:
        column_map["jlac10_standard_name"] = args.col_std_name

    if "item_name" not in column_map or "jlac10" not in column_map:
        print(
            "エラー: --col-item と --col-jlac10 は必須です",
            file=sys.stderr,
        )
        return 1

    output_path = Path(args.output_file) if args.output_file else None

    try:
        result = convert_tabular(
            filepath=filepath,
            column_map=column_map,
            hospital=args.hospital,
            sheet_name=args.sheet,
            skip_rows=args.skip_rows,
            output_path=output_path,
        )
    except (FileNotFoundError, ValueError) as e:
        print(f"エラー: {e}", file=sys.stderr)
        return 1

    meta = result["metadata"]
    out = output_path or filepath.with_suffix(".json")
    print(f"\n変換完了: {meta['total_items']}件")
    print(f"  病院: {meta['hospital'] or '(未指定)'}")
    print(f"  ソース: {meta['source_file']}")
    print(f"  出力: {out}")
    return 0


# ---------------------------------------------------------------------------
# Map (院内項目 → JLAC10 マッピング)
# ---------------------------------------------------------------------------

def cmd_map(args: argparse.Namespace) -> int:
    filepath = Path(args.input)

    # converter の convert_tabular で入力を読む（item_name列のみ必須）
    column_map: dict[str, str] = {"item_name": args.col_name}
    # jlac10列は任意（ダミー列としてitem_nameと同じにしておく）
    if args.col_jlac10:
        column_map["jlac10"] = args.col_jlac10
    else:
        column_map["jlac10"] = args.col_name  # ダミー: convert_tabularの必須要件を満たす

    try:
        converted = convert_tabular(
            filepath=filepath,
            column_map=column_map,
            hospital=args.hospital,
            sheet_name=args.sheet,
            skip_rows=args.skip_rows,
            output_path=filepath.with_suffix(".tmp.json"),
        )
    except (FileNotFoundError, ValueError) as e:
        print(f"エラー: {e}", file=sys.stderr)
        return 1
    finally:
        # 一時JSONを削除
        tmp = filepath.with_suffix(".tmp.json")
        if tmp.exists():
            tmp.unlink()

    items = converted["items"]
    if not items:
        print("エラー: 入力データが空です", file=sys.stderr)
        return 1

    # 検索インデックス構築
    data_dir = Path(args.data_dir)
    try:
        index = build_index(data_dir)
    except FileNotFoundError as e:
        print(f"エラー: {e}", file=sys.stderr)
        return 1

    print(f"インデックス構築完了: {len(index.entries)}件")
    print(f"入力項目数: {len(items)}件")

    # 一括マッピング実行
    results = bulk_map(
        items=items,
        index=index,
        auto_threshold=args.threshold_auto,
        candidate_threshold=args.threshold_candidate,
        max_candidates=args.max_candidates,
    )

    # Excel出力
    output_path = Path(args.output)
    export_mapping_excel(results, output_path)

    # JSON出力
    json_path = output_path.with_suffix(".json")
    export_mapping_json(results, json_path)

    meta = results["metadata"]
    print(f"\nマッピング完了: {meta['total']}件")
    print(f"  自動マッピング (auto):     {meta['auto']}件")
    print(f"  候補あり (candidate):      {meta['candidate']}件")
    print(f"  手動確認必要 (manual):     {meta['manual']}件")
    print(f"  Excel出力: {output_path}")
    print(f"  JSON出力:  {json_path}")
    return 0


def cmd_map_auto(args: argparse.Namespace) -> int:
    """ベンダー自動検出で一括マッピング + 元Excel追記"""
    input_path = Path(args.input)
    data_dir = Path(args.data_dir)
    vendor = args.vendor
    hospital = args.hospital or ""
    sheet = args.sheet

    # 1. 自動変換
    logger.info("自動変換: vendor=%s, sheet=%s", vendor, sheet)
    converted = convert_auto(
        filepath=input_path,
        vendor=vendor,
        hospital=hospital,
        sheet_name=sheet,
        skip_rows=args.skip_rows,
    )
    items = converted["items"]
    print(f"\n変換完了: {len(items)}件")

    # 2. マッピング
    index = build_index(data_dir)
    results = bulk_map(
        items=items,
        index=index,
        auto_threshold=args.threshold_auto,
        candidate_threshold=args.threshold_candidate,
    )
    meta = results["metadata"]
    print(f"マッピング: auto={meta['auto']} / candidate={meta['candidate']} / manual={meta['manual']}")

    # 3. 元Excelに追記
    output_path = Path(args.output) if args.output else input_path.with_name(
        input_path.stem + "_mapped" + input_path.suffix
    )
    write_jlac10_to_excel(
        source_path=input_path,
        mapping_results=results["results"],
        output_path=output_path,
        sheet_name=sheet,
        skip_rows=args.skip_rows,
    )
    print(f"出力: {output_path}")

    # 4. JSON も出力
    json_path = output_path.with_suffix(".json")
    from .mapper import export_mapping_json
    export_mapping_json(results, json_path)

    return 0


# ---------------------------------------------------------------------------
# DB還元
# ---------------------------------------------------------------------------

def cmd_apply_mapping(args: argparse.Namespace) -> int:
    """確定済みマッピング結果をDBに還元"""
    input_path = Path(args.input)
    data_dir = Path(args.data_dir)
    merged_path = data_dir / "merged_jlac10.json"
    hospital = args.hospital or "Unknown"

    if not merged_path.exists():
        print(f"merged_jlac10.json が見つかりません: {merged_path}", file=sys.stderr)
        return 1

    # 入力がJSON
    if input_path.suffix.lower() == ".json":
        data = json.loads(input_path.read_text(encoding="utf-8"))
        items = data.get("results", data.get("items", []))
    # 入力がExcel（マッピング結果Excel）
    else:
        import openpyxl
        wb = openpyxl.load_workbook(str(input_path), read_only=True)
        ws = wb[wb.sheetnames[0]]
        items = []
        headers = None
        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = [str(c or "").lower() for c in row]
                continue
            row_dict = dict(zip(headers, row))
            status = str(row_dict.get("status", "")).lower()
            if status in ("auto", "confirmed", "ok"):
                items.append({
                    "status": "confirmed",
                    "item_name": str(row_dict.get("item name", row_dict.get("item_name", ""))),
                    "jlac10": str(row_dict.get("jlac10", row_dict.get("matched jlac10", ""))),
                    "matched_name": str(row_dict.get("matched name", row_dict.get("matched_name", ""))),
                })
        wb.close()

    result = apply_mapping_results(
        merged_path=merged_path,
        mapping_items=items,
        hospital=hospital,
        confirmed_only=args.confirmed_only,
    )

    print(f"\nDB還元完了:")
    print(f"  追加: {result['added']}件")
    print(f"  スキップ: {result['skipped']}件")
    print(f"  新規エントリ: {result['new_entries']}件")
    return 0


# ---------------------------------------------------------------------------
# NCDA差異チェック
# ---------------------------------------------------------------------------

def cmd_check_ncda(args: argparse.Namespace) -> int:
    """外注先JLAC10 vs NCDA 差異チェック"""
    from .converter import _read_xlsx, _read_csv, _resolve_column_index
    from .merge import load_jlac10_lookup

    filepath = Path(args.input)
    if not filepath.exists():
        print(f"ファイルが見つかりません: {filepath}", file=sys.stderr)
        return 1

    # データ読み込み
    suffix = filepath.suffix.lower()
    if suffix == ".xlsx":
        data_rows, header_row = _read_xlsx(filepath, args.sheet, args.skip_rows)
    elif suffix == ".csv":
        data_rows, header_row = _read_csv(filepath, args.skip_rows)
    else:
        print(f"未対応のファイル形式: {suffix} (.xlsx / .csv)", file=sys.stderr)
        return 1

    # 列指定を解決
    outsource_col = _resolve_column_index(args.outsource_col, header_row)
    ncda_col = _resolve_column_index(args.ncda_col, header_row)
    name_col = _resolve_column_index(args.name_col, header_row) if args.name_col else None

    # lookup 読み込み
    data_dir = Path(args.data_dir)
    lookup = load_jlac10_lookup(data_dir)
    if not lookup:
        print(
            "警告: jlac10_lookup.json が見つかりません。"
            "コード名称は表示されません。",
            file=sys.stderr,
        )

    # チェック用データ構築
    items: list[dict] = []
    for row in data_rows:
        if not row or all(cell.strip() == "" for cell in row):
            continue
        outsource = row[outsource_col].strip() if outsource_col < len(row) else ""
        ncda = row[ncda_col].strip() if ncda_col < len(row) else ""
        item_name = ""
        if name_col is not None and name_col < len(row):
            item_name = row[name_col].strip()

        if not outsource and not ncda:
            continue

        items.append({
            "outsource_jlac10": outsource,
            "ncda_jlac10": ncda,
            "item_name": item_name,
        })

    if not items:
        print("エラー: チェック対象データが空です", file=sys.stderr)
        return 1

    # 一括チェック実行
    results = ncda_batch_check(items, lookup)
    meta = results["metadata"]

    # Excel出力
    output_path = Path(args.output)
    export_check_excel(results, output_path)

    print(f"\nNCDA差異チェック完了: {meta['total']}件")
    print(f"  OK:      {meta['ok']}件")
    print(f"  Warning: {meta['warnings']}件")
    print(f"  Error:   {meta['errors']}件")
    print(f"  出力: {output_path}")
    return 0


# ---------------------------------------------------------------------------
# 統合・差分・チェック・一覧
# ---------------------------------------------------------------------------

def cmd_merge(args: argparse.Namespace) -> int:
    result = merge_all(output_dir=Path(args.output))
    meta = result["metadata"]
    print(f"\n統合完了: {meta['total_unique_jlac10']}件のユニーク JLAC10")
    print(f"  データソース: {meta['sources_available']}")
    counts = meta["by_source_count"]
    print(f"  3社共通: {counts['all_three']}件 / 2社: {counts['two_sources']}件")
    print(f"  SRLのみ: {counts['srl_only']}件 / BMLのみ: {counts['bml_only']}件 / LSIのみ: {counts['lsi_only']}件")
    return 0


def cmd_check(args: argparse.Namespace) -> int:
    output_dir = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)
    session = requests.Session()

    import time
    from .scraper import REQUEST_INTERVAL

    # SRL
    srl_needed, srl_remote, srl_local = check_update_needed(output_dir, session)
    print(f"[SRL] Last Up Date: {srl_remote} (前回: {srl_local or '未取得'})")
    print(f"  → {'更新あり' if srl_needed else '更新なし'}")
    time.sleep(REQUEST_INTERVAL)

    # BML
    bml_needed, bml_remote, bml_local = check_bml_update_needed(output_dir, session)
    print(f"[BML] 最新掲載日:  {bml_remote} (前回: {bml_local or '未取得'})")
    print(f"  → {'更新あり' if bml_needed else '更新なし'}")
    time.sleep(REQUEST_INTERVAL)

    # LSI
    lsi_needed, lsi_remote, lsi_local = check_lsi_update_needed(output_dir, session)
    print(f"[LSI] 掲載日:      {lsi_remote} (前回: {lsi_local or '未取得'})")
    print(f"  → {'更新あり' if lsi_needed else '更新なし'}")
    time.sleep(REQUEST_INTERVAL)

    # JSLM
    jslm_needed, jslm_remote, jslm_local = check_jslm_update_needed(output_dir, session)
    print(f"[JSLM] コード表:   version {jslm_remote} (前回: {jslm_local or '未取得'})")
    print(f"  → {'更新あり' if jslm_needed else '更新なし'}")

    return 0


def cmd_diff(args: argparse.Namespace) -> int:
    old_path = Path(args.old)
    new_path = Path(args.new)

    if not old_path.exists():
        print(f"ファイルが見つかりません: {old_path}", file=sys.stderr)
        return 1
    if not new_path.exists():
        print(f"ファイルが見つかりません: {new_path}", file=sys.stderr)
        return 1

    report = diff_report(old_path, new_path)
    s = report["summary"]
    print(f"\n差分レポート:")
    print(f"  追加: {s['added']}件 / 削除: {s['removed']}件 / 変更: {s['changed']}件 / 変更なし: {s['unchanged']}件")

    if args.output_report:
        out = Path(args.output_report)
        out.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"  レポート保存: {out}")
    return 0


def cmd_search(args: argparse.Namespace) -> int:
    data_dir = Path(args.output)
    index = build_index(data_dir)
    print(f"インデックス構築完了: {len(index.entries)}件")

    if args.query:
        # 引数で検索
        results = index.search(args.query, max_results=args.max)
        print(f"\n検索: 「{args.query}」 → {len(results)}件ヒット")
        print(format_results(results))
    else:
        # 対話モード
        print("検索クエリを入力してください（quit で終了）:\n")
        while True:
            try:
                query = input("検索> ").strip()
            except (EOFError, KeyboardInterrupt):
                print()
                break
            if not query or query.lower() in ("quit", "exit", "q"):
                break
            results = index.search(query, max_results=args.max)
            print(f"\n「{query}」 → {len(results)}件ヒット")
            print(format_results(results))
            print()
    return 0


def cmd_list(args: argparse.Namespace) -> int:
    current_group = ""
    for cat in CATEGORIES:
        if cat.group != current_group:
            current_group = cat.group
            print(f"\n[{current_group}]")
        print(f"  {cat.id:>3d}: {cat.name}")
    return 0


# ---------------------------------------------------------------------------
# メイン
# ---------------------------------------------------------------------------

def _add_cache_args(parser: argparse.ArgumentParser) -> None:
    parser.add_argument("--no-cache", action="store_true", help="キャッシュ無視")
    parser.add_argument("--cache-age", type=float, default=24.0, help="キャッシュ有効期限(時間)")
    parser.add_argument("-o", "--output", default="data", help="出力ディレクトリ")


def main() -> None:
    parser = argparse.ArgumentParser(
        prog="srl-scraper",
        description="外注検査 JLAC10 データ取得ツール（SRL / BML / LSI メディエンス）",
    )
    parser.add_argument("-v", "--verbose", action="store_true", help="詳細ログ出力")
    sub = parser.add_subparsers(dest="command")

    # srl
    p_srl = sub.add_parser("srl", help="SRL検査案内からデータ取得")
    _add_cache_args(p_srl)
    p_srl.add_argument("-c", "--categories", help="カテゴリIDをカンマ区切り (例: 1,2,3)")
    p_srl.add_argument("--check-update", action="store_true", help="Last Up Date 確認、更新時のみ取得")
    p_srl.add_argument("--force", action="store_true", help="check-update時でも強制取得")

    # bml
    p_bml = sub.add_parser("bml", help="BML検査案内からデータ取得")
    _add_cache_args(p_bml)
    p_bml.add_argument("--check-update", action="store_true", help="new/changeの掲載日を確認、更新時のみ取得")

    # lsi
    p_lsi = sub.add_parser("lsi", help="LSIメディエンス検査案内からデータ取得")
    _add_cache_args(p_lsi)
    p_lsi.add_argument("--check-update", action="store_true", help="掲載日を確認、更新時のみ取得")

    # jslm
    p_jslm = sub.add_parser("jslm", help="JSLM JLAC10コード表を取得")
    p_jslm.add_argument("-o", "--output", default="data", help="出力ディレクトリ")
    p_jslm.add_argument("--check-update", action="store_true", help="版数確認、更新時のみ取得")

    # reagent
    p_reagent = sub.add_parser("reagent", help="試薬DB構築（メーカー製品 + PMDA添付文書）")
    p_reagent.add_argument("-o", "--output", default="data", help="出力ディレクトリ")
    p_reagent.add_argument("--pmda", help="PMDA添付文書URLを追加")

    # sop
    p_sop = sub.add_parser("sop", help="SOP(Word/PDF)から測定法・試薬情報を抽出")
    p_sop.add_argument("path", help="SOPファイル(.docx/.pdf) または ディレクトリ")
    p_sop.add_argument("-o", "--output", default="data", help="出力ディレクトリ")

    # search
    p_search = sub.add_parser("search", help="検査項目をあいまい検索")
    p_search.add_argument("query", nargs="?", default=None, help="検索文字列（省略で対話モード）")
    p_search.add_argument("-o", "--output", default="data", help="データディレクトリ")
    p_search.add_argument("-n", "--max", type=int, default=10, help="最大結果数 (default: 10)")

    # apply-mapping
    p_apply = sub.add_parser("apply-mapping", help="マッピング結果をDBに還元")
    p_apply.add_argument("input", help="マッピング結果 (.json / .xlsx)")
    p_apply.add_argument("-d", "--data-dir", default="data", help="データディレクトリ")
    p_apply.add_argument("--hospital", default="", help="病院名")
    p_apply.add_argument("--confirmed-only", action="store_true", default=True, help="確定分のみ適用（デフォルト）")
    p_apply.add_argument("--all", dest="confirmed_only", action="store_false", help="全項目を適用")

    # merge
    p_merge = sub.add_parser("merge", help="3社のデータをJLAC10で統合")
    p_merge.add_argument("-o", "--output", default="data", help="出力ディレクトリ")

    # check
    p_check = sub.add_parser("check", help="全ソースの更新状況を確認")
    p_check.add_argument("-o", "--output", default="data", help="出力ディレクトリ")

    # diff
    p_diff = sub.add_parser("diff", help="2つのJSONファイルの差分表示")
    p_diff.add_argument("old", help="古いJSON")
    p_diff.add_argument("new", help="新しいJSON")
    p_diff.add_argument("-o", "--output-report", help="差分レポート出力先")

    # convert
    p_convert = sub.add_parser("convert", help="院内検査マスタ(Excel/CSV)をJSONに変換")
    p_convert.add_argument("input", help="入力ファイル (.xlsx / .csv)")
    p_convert.add_argument("-o", "--output-file", default=None, help="出力JSONパス (省略で {入力ファイル名}.json)")
    p_convert.add_argument("--hospital", default="", help="病院名")
    p_convert.add_argument("--col-item", required=True, help="検査項目名の列 (A, 1, またはヘッダ名)")
    p_convert.add_argument("--col-jlac10", required=True, help="JLAC10の列 (A, 1, またはヘッダ名)")
    p_convert.add_argument("--col-abbr", default=None, help="略称の列 (A, 1, またはヘッダ名)")
    p_convert.add_argument("--col-std-name", default=None, help="JLAC10標準名称の列 (A, 1, またはヘッダ名)")
    p_convert.add_argument("--sheet", default=None, help="Excelシート名 (省略で最初のシート)")
    p_convert.add_argument("--skip-rows", type=int, default=1, help="スキップするヘッダ行数 (default: 1)")

    # map
    p_map = sub.add_parser("map", help="院内検査項目をJLAC10にマッピング")
    p_map.add_argument("input", help="入力ファイル (.xlsx / .csv)")
    p_map.add_argument("--col-name", required=True, help="検査項目名の列 (A, 1, またはヘッダ名)")
    p_map.add_argument("--col-jlac10", default=None, help="既存JLAC10の列 (参考用)")
    p_map.add_argument("-o", "--output", default="mapping_result.xlsx", help="出力Excelパス (default: mapping_result.xlsx)")
    p_map.add_argument("-d", "--data-dir", default="data", help="データディレクトリ (default: data)")
    p_map.add_argument("--hospital", default="", help="病院名")
    p_map.add_argument("--sheet", default=None, help="Excelシート名 (省略で最初のシート)")
    p_map.add_argument("--skip-rows", type=int, default=1, help="スキップするヘッダ行数 (default: 1)")
    p_map.add_argument("--threshold-auto", type=float, default=90.0, help="自動マッピング閾値 (default: 90)")
    p_map.add_argument("--threshold-candidate", type=float, default=50.0, help="候補閾値 (default: 50)")
    p_map.add_argument("--max-candidates", type=int, default=5, help="候補最大件数 (default: 5)")

    # map-auto
    p_mauto = sub.add_parser("map-auto", help="ベンダー自動検出で一括マッピング + 元Excel追記")
    p_mauto.add_argument("input", help="入力ファイル (.xlsx)")
    p_mauto.add_argument("--vendor", default=None, help="ベンダー名 (NEC/Fujitsu/IBM/SSI/SBS/KHI/CSI/NAIS)")
    p_mauto.add_argument("-o", "--output", default=None, help="出力先 (省略で {入力}_mapped.xlsx)")
    p_mauto.add_argument("-d", "--data-dir", default="data", help="データディレクトリ")
    p_mauto.add_argument("--hospital", default="", help="病院名")
    p_mauto.add_argument("--sheet", default=None, help="シート名")
    p_mauto.add_argument("--skip-rows", type=int, default=1, help="ヘッダ行数")
    p_mauto.add_argument("--threshold-auto", type=float, default=90.0, help="自動マッピング閾値")
    p_mauto.add_argument("--threshold-candidate", type=float, default=50.0, help="候補閾値")

    # vendors
    sub.add_parser("vendors", help="登録済みベンダー一覧")

    # check-ncda
    p_ncda = sub.add_parser("check-ncda", help="外注先JLAC10 vs NCDA 差異チェック")
    p_ncda.add_argument("input", help="入力ファイル (.xlsx / .csv)")
    p_ncda.add_argument("--outsource-col", required=True, help="外注先JLAC10の列 (A, 1, またはヘッダ名)")
    p_ncda.add_argument("--ncda-col", required=True, help="NCDA JLAC10の列 (A, 1, またはヘッダ名)")
    p_ncda.add_argument("--name-col", default=None, help="検査項目名の列 (A, 1, またはヘッダ名)")
    p_ncda.add_argument("-o", "--output", default="ncda_report.xlsx", help="出力Excelパス (default: ncda_report.xlsx)")
    p_ncda.add_argument("-d", "--data-dir", default="data", help="データディレクトリ (default: data)")
    p_ncda.add_argument("--sheet", default=None, help="Excelシート名 (省略で最初のシート)")
    p_ncda.add_argument("--skip-rows", type=int, default=1, help="スキップするヘッダ行数 (default: 1)")

    # list
    sub.add_parser("list", help="SRLカテゴリ一覧")

    args = parser.parse_args()
    setup_logging(args.verbose)

    if not args.command:
        parser.print_help()
        sys.exit(1)

    commands = {
        "srl": cmd_srl,
        "bml": cmd_bml,
        "lsi": cmd_lsi,
        "jslm": cmd_jslm,
        "reagent": cmd_reagent,
        "sop": cmd_sop,
        "search": cmd_search,
        "apply-mapping": cmd_apply_mapping,
        "merge": cmd_merge,
        "check": cmd_check,
        "diff": cmd_diff,
        "convert": cmd_convert,
        "map": cmd_map,
        "map-auto": cmd_map_auto,
        "vendors": lambda args: (print("\n".join(list_vendors())), 0)[1],
        "check-ncda": cmd_check_ncda,
        "list": cmd_list,
    }
    sys.exit(commands[args.command](args))
