"""外注先JLAC10(15桁) vs 院内NCDA(17桁) 差異チェックモジュール"""

import logging
from pathlib import Path

from .scraper import classify_jlac10

logger = logging.getLogger(__name__)

# JLAC10構造定義
# 分析物(5) + 識別(4) + 材料(3) + 測定法(3) + 結果識別(2、NCDAのみ)
JLAC10_FIELDS = [
    {"name": "analyte", "label": "分析物", "start": 0, "end": 5, "lookup_key": "analyte"},
    {"name": "identification", "label": "識別", "start": 5, "end": 9, "lookup_key": "identification"},
    {"name": "material", "label": "材料", "start": 9, "end": 12, "lookup_key": "material"},
    {"name": "method", "label": "測定法", "start": 12, "end": 15, "lookup_key": "method"},
]

# フィールドごとの severity ルール
FIELD_SEVERITY = {
    "analyte": "error",
    "identification": "error",
    "material": "error",
    "method": "warning",
}


# ---------------------------------------------------------------------------
# コード分解ユーティリティ
# ---------------------------------------------------------------------------

def _split_jlac10(code: str) -> dict[str, str]:
    """JLAC10コードを各パートに分解する。

    15桁以上のコードを想定。15桁未満の場合は取得可能な範囲で分解。
    """
    code = code.replace("-", "")
    parts: dict[str, str] = {}
    for field in JLAC10_FIELDS:
        if len(code) >= field["end"]:
            parts[field["name"]] = code[field["start"]:field["end"]]
        else:
            parts[field["name"]] = ""
    # 結果識別 (16-17桁目、17桁コードのみ)
    if len(code) >= 17:
        parts["result_identification"] = code[15:17]
    elif len(code) == 16:
        parts["result_identification"] = code[15:16]
    else:
        parts["result_identification"] = ""
    return parts


def _lookup_name(lookup: dict, lookup_key: str, code: str) -> str:
    """lookup辞書からコードの名称を取得する。"""
    section = lookup.get(lookup_key, {})
    entry = section.get(code, {})
    return entry.get("name", "")


# ---------------------------------------------------------------------------
# 差異チェック
# ---------------------------------------------------------------------------

def check_outsource_vs_ncda(
    outsource_jlac10: str,
    ncda_jlac10: str,
    lookup: dict,
) -> list[dict]:
    """外注先JLAC10(15桁)とNCDA(17桁)の各パートを比較して差異を検出する。

    Args:
        outsource_jlac10: 外注先提出の15桁JLAC10コード
        ncda_jlac10: 院内NCDA 17桁コード
        lookup: jlac10_lookup.json から読み込んだ辞書

    Returns:
        差異情報のリスト。差異がなければ空リスト。
        各要素: {
            "field": "analyte|identification|material|method",
            "outsource_code": str,
            "ncda_code": str,
            "outsource_name": str,
            "ncda_name": str,
            "severity": "error|warning"
        }
    """
    outsource_parts = _split_jlac10(outsource_jlac10)
    ncda_parts = _split_jlac10(ncda_jlac10)

    diffs: list[dict] = []
    for field in JLAC10_FIELDS:
        name = field["name"]
        lookup_key = field["lookup_key"]
        o_code = outsource_parts.get(name, "")
        n_code = ncda_parts.get(name, "")

        if o_code != n_code:
            diffs.append({
                "field": name,
                "outsource_code": o_code,
                "ncda_code": n_code,
                "outsource_name": _lookup_name(lookup, lookup_key, o_code),
                "ncda_name": _lookup_name(lookup, lookup_key, n_code),
                "severity": FIELD_SEVERITY.get(name, "warning"),
            })

    return diffs


def check_result_identification(
    ncda_jlac10: str,
    lookup: dict,
) -> dict:
    """結果識別コード(16-17桁目)の妥当性チェック。

    result_common (共通コード) に存在するか確認する。

    Args:
        ncda_jlac10: 17桁NCDAコード
        lookup: jlac10_lookup.json から読み込んだ辞書

    Returns:
        {"code": "XX", "valid": True/False, "name": "...", "message": "..."}
    """
    parts = _split_jlac10(ncda_jlac10)
    code = parts.get("result_identification", "")

    if not code:
        return {
            "code": "",
            "valid": False,
            "name": "",
            "message": "結果識別コードが取得できません（コード長不足）",
        }

    result_common = lookup.get("result_common", {})
    # result_common のキーは 3桁 zero-padded の場合がある
    # 2桁コードの先頭に 0 を付けて検索も試みる
    entry = result_common.get(code) or result_common.get(code.zfill(3))

    if entry:
        return {
            "code": code,
            "valid": True,
            "name": entry.get("name", ""),
            "message": f"結果識別コード '{code}' は有効です: {entry.get('name', '')}",
        }

    return {
        "code": code,
        "valid": False,
        "name": "",
        "message": f"結果識別コード '{code}' は result_common に存在しません",
    }


# ---------------------------------------------------------------------------
# 一括チェック
# ---------------------------------------------------------------------------

def batch_check(
    items: list[dict],
    lookup: dict,
) -> dict:
    """一括差異チェックを実行する。

    Args:
        items: [{"outsource_jlac10": str, "ncda_jlac10": str, "item_name": str}, ...]
        lookup: jlac10_lookup.json から読み込んだ辞書

    Returns:
        {
            "metadata": {"total": N, "errors": N, "warnings": N, "ok": N},
            "results": [
                {
                    "item_name": str,
                    "outsource": str,
                    "ncda": str,
                    "outsource_status": str,
                    "ncda_status": str,
                    "diffs": [...],
                    "result_id_check": {...},
                    "status": "ok|warning|error"
                }
            ]
        }
    """
    results: list[dict] = []
    counts = {"errors": 0, "warnings": 0, "ok": 0}

    for item in items:
        outsource = item.get("outsource_jlac10", "").replace("-", "")
        ncda = item.get("ncda_jlac10", "").replace("-", "")
        item_name = item.get("item_name", "")

        outsource_status = classify_jlac10(outsource)
        ncda_status = classify_jlac10(ncda)

        # コード形式チェック
        diffs: list[dict] = []
        result_id_check: dict = {}

        if outsource_status in ("valid_15", "valid_17") and ncda_status in ("valid_15", "valid_17"):
            diffs = check_outsource_vs_ncda(outsource, ncda, lookup)
            if ncda_status == "valid_17":
                result_id_check = check_result_identification(ncda, lookup)
        elif outsource_status in ("empty", "invalid") or ncda_status in ("empty", "invalid"):
            # コード不正の場合はエラー扱いの差異として記録
            if outsource_status in ("empty", "invalid"):
                diffs.append({
                    "field": "outsource_format",
                    "outsource_code": outsource,
                    "ncda_code": ncda,
                    "outsource_name": "",
                    "ncda_name": "",
                    "severity": "error",
                })
            if ncda_status in ("empty", "invalid"):
                diffs.append({
                    "field": "ncda_format",
                    "outsource_code": outsource,
                    "ncda_code": ncda,
                    "outsource_name": "",
                    "ncda_name": "",
                    "severity": "error",
                })

        # ステータス判定
        has_error = any(d["severity"] == "error" for d in diffs)
        has_warning = any(d["severity"] == "warning" for d in diffs)
        ri_invalid = result_id_check.get("valid") is False if result_id_check else False

        if has_error or ri_invalid:
            status = "error"
            counts["errors"] += 1
        elif has_warning:
            status = "warning"
            counts["warnings"] += 1
        else:
            status = "ok"
            counts["ok"] += 1

        results.append({
            "item_name": item_name,
            "outsource": outsource,
            "ncda": ncda,
            "outsource_status": outsource_status,
            "ncda_status": ncda_status,
            "diffs": diffs,
            "result_id_check": result_id_check,
            "status": status,
        })

    return {
        "metadata": {
            "total": len(results),
            "errors": counts["errors"],
            "warnings": counts["warnings"],
            "ok": counts["ok"],
        },
        "results": results,
    }


# ---------------------------------------------------------------------------
# Excel出力
# ---------------------------------------------------------------------------

def export_check_excel(results: dict, output_path: Path) -> None:
    """差異チェック結果を色分けExcelに出力する。

    行の色分け:
        ok      = 緑背景 (C6EFCE)
        warning = 黄背景 (FFEB9C)
        error   = 赤背景 (FFC7CE)
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NCDA Check Results"

    # ヘッダー定義
    headers = [
        "Status",
        "Item Name",
        "Outsource JLAC10",
        "NCDA JLAC10",
        "Diff Fields",
        "Diff Details",
        "Result ID Code",
        "Result ID Valid",
        "Result ID Name",
    ]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 行の色分け
    fill_ok = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_warning = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    fill_error = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    status_fill_map = {
        "ok": fill_ok,
        "warning": fill_warning,
        "error": fill_error,
    }

    for row_idx, entry in enumerate(results["results"], start=2):
        status = entry["status"]
        diffs = entry.get("diffs", [])
        ri = entry.get("result_id_check", {})

        # 差異フィールド名を連結
        diff_fields = ", ".join(d["field"] for d in diffs) if diffs else ""
        # 差異詳細を構築
        diff_details_parts = []
        for d in diffs:
            o_label = f"{d['outsource_code']}"
            if d["outsource_name"]:
                o_label += f"({d['outsource_name']})"
            n_label = f"{d['ncda_code']}"
            if d["ncda_name"]:
                n_label += f"({d['ncda_name']})"
            diff_details_parts.append(
                f"[{d['field']}] {o_label} -> {n_label} ({d['severity']})"
            )
        diff_details = "\n".join(diff_details_parts)

        values = [
            status.upper(),
            entry["item_name"],
            entry["outsource"],
            entry["ncda"],
            diff_fields,
            diff_details,
            ri.get("code", ""),
            "OK" if ri.get("valid") else ("NG" if ri.get("valid") is False else ""),
            ri.get("name", ""),
        ]

        fill = status_fill_map.get(status)
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if fill:
                cell.fill = fill

    # 列幅調整
    column_widths = [10, 30, 20, 22, 25, 50, 12, 10, 20]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = width

    # サマリーシート
    ws_summary = wb.create_sheet("Summary", 0)
    meta = results["metadata"]
    summary_data = [
        ["NCDA差異チェック サマリー", ""],
        ["", ""],
        ["総件数", meta["total"]],
        ["OK", meta["ok"]],
        ["Warning", meta["warnings"]],
        ["Error", meta["errors"]],
    ]
    for row_idx, (label, value) in enumerate(summary_data, start=1):
        ws_summary.cell(row=row_idx, column=1, value=label)
        ws_summary.cell(row=row_idx, column=2, value=value)
    ws_summary.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 15

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info("NCDAチェック結果出力: %s", output_path)
