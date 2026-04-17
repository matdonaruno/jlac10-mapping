"""院内検査マスタ（Excel/CSV）を統一JSONフォーマットに変換するモジュール

ベンダー別のヘッダー構造に対応:
  - ベンダー名指定で自動列検出
  - ベンダー不明でも汎用キーワードで推定
  - 手動列指定も引き続きサポート
"""

import csv
import json
import logging
import re
from datetime import datetime, timezone
from pathlib import Path

from .scraper import classify_jlac10
from .vendor_profiles import detect_columns

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# 列指定のパース
# ---------------------------------------------------------------------------

def _parse_column_spec(spec: str) -> int | str:
    """列指定文字列をパースする。

    対応形式:
      - アルファベット列名: "A", "B", "AA" → 0-indexed 整数
      - 数値文字列: "1", "2", "3" → 0-indexed 整数 (1始まりを0始まりに変換)
      - ヘッダ名: "検査項目名称" などの文字列 → そのまま返す

    Returns:
        int (0-indexed列番号) または str (ヘッダ名)
    """
    spec = spec.strip()

    # 数値指定 ("1", "2", "3" ...)
    if spec.isdigit():
        idx = int(spec) - 1
        if idx < 0:
            raise ValueError(f"列番号は1以上を指定してください: {spec}")
        return idx

    # アルファベット列名 ("A", "B", "AA" ...)
    if re.fullmatch(r"[A-Za-z]{1,3}", spec):
        result = 0
        for ch in spec.upper():
            result = result * 26 + (ord(ch) - ord("A") + 1)
        return result - 1  # 0-indexed

    # ヘッダ名（それ以外の文字列）
    return spec


def _resolve_column_index(spec: str, headers: list[str] | None) -> int:
    """列指定を確定した0-indexedの列番号に変換する。

    ヘッダ名指定の場合は headers が必要。
    """
    parsed = _parse_column_spec(spec)
    if isinstance(parsed, int):
        return parsed

    # ヘッダ名で検索
    header_name = parsed
    if headers is None:
        raise ValueError(
            f"ヘッダ名 '{header_name}' で列を指定していますが、"
            "ヘッダ行が見つかりません"
        )
    # 完全一致
    for i, h in enumerate(headers):
        if h.strip() == header_name:
            return i
    # 部分一致（フォールバック）
    for i, h in enumerate(headers):
        if header_name in h.strip():
            logger.warning(
                "列 '%s' を部分一致で解決しました: '%s' (列%d)",
                header_name, h.strip(), i + 1,
            )
            return i
    raise ValueError(
        f"ヘッダ名 '{header_name}' に一致する列が見つかりません。"
        f" 利用可能なヘッダ: {headers}"
    )


# ---------------------------------------------------------------------------
# Excel (.xlsx) 読み込み
# ---------------------------------------------------------------------------

def _read_xlsx(
    filepath: Path,
    sheet_name: str | None,
    skip_rows: int,
) -> tuple[list[list[str]], list[str] | None]:
    """openpyxl で Excel を読み込み、文字列の2次元リストとヘッダを返す。

    Returns:
        (data_rows, header_row)
        skip_rows > 0 の場合、先頭 skip_rows 行のうち最後の行をヘッダとして返す。
    """
    import openpyxl

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    try:
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise ValueError(
                    f"シート '{sheet_name}' が見つかりません。"
                    f" 利用可能: {wb.sheetnames}"
                )
            ws = wb[sheet_name]
        else:
            ws = wb.active

        all_rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            all_rows.append([str(cell) if cell is not None else "" for cell in row])
    finally:
        wb.close()

    header_row = None
    if skip_rows > 0 and len(all_rows) >= skip_rows:
        header_row = all_rows[skip_rows - 1]
        data_rows = all_rows[skip_rows:]
    else:
        data_rows = all_rows

    return data_rows, header_row


# ---------------------------------------------------------------------------
# CSV (.csv) 読み込み
# ---------------------------------------------------------------------------

def _read_csv(
    filepath: Path,
    skip_rows: int,
) -> tuple[list[list[str]], list[str] | None]:
    """CSV を読み込み、文字列の2次元リストとヘッダを返す。"""
    all_rows: list[list[str]] = []

    # エンコーディングを自動検出
    for encoding in ("utf-8-sig", "utf-8", "cp932", "shift_jis"):
        try:
            with filepath.open("r", encoding=encoding, newline="") as f:
                reader = csv.reader(f)
                all_rows = [row for row in reader]
            logger.debug("CSVエンコーディング検出: %s (%d行)", encoding, len(all_rows))
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError(
            f"ファイル '{filepath}' のエンコーディングを判定できません。"
            " UTF-8, CP932, Shift_JIS を試しました。"
        )

    header_row = None
    if skip_rows > 0 and len(all_rows) >= skip_rows:
        header_row = all_rows[skip_rows - 1]
        data_rows = all_rows[skip_rows:]
    else:
        data_rows = all_rows

    return data_rows, header_row


# ---------------------------------------------------------------------------
# メイン変換関数
# ---------------------------------------------------------------------------

def convert_tabular(
    filepath: Path,
    column_map: dict[str, str],
    hospital: str = "",
    sheet_name: str | None = None,
    skip_rows: int = 1,
    output_path: Path | None = None,
) -> dict:
    """院内検査マスタ (Excel/CSV) を統一JSON形式に変換する。

    Args:
        filepath: 入力ファイルパス (.xlsx または .csv)
        column_map: フィールド名→列指定のマッピング
            必須キー: "item_name", "jlac10"
            任意キー: "abbreviation", "jlac10_standard_name"
            列指定: "A"/"B" (アルファベット), "1"/"2" (数値), "検査名" (ヘッダ名)
        hospital: 病院名
        sheet_name: Excelシート名 (省略で最初のシート)
        skip_rows: スキップするヘッダ行数 (デフォルト1)
        output_path: 出力JSONパス (省略で {入力ファイル名}.json)

    Returns:
        出力JSONと同じ構造の dict
    """
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"ファイルが見つかりません: {filepath}")

    suffix = filepath.suffix.lower()
    logger.debug("ファイル読み込み開始: %s (形式: %s)", filepath.name, suffix)
    if suffix == ".xlsx":
        data_rows, header_row = _read_xlsx(filepath, sheet_name, skip_rows)
    elif suffix == ".csv":
        data_rows, header_row = _read_csv(filepath, skip_rows)
    else:
        raise ValueError(
            f"未対応のファイル形式: {suffix} (.xlsx または .csv に対応)"
        )
    logger.debug(
        "ファイル読み込み完了: %s (データ行数: %d, ヘッダ: %s)",
        filepath.name, len(data_rows),
        header_row[:5] if header_row else None,
    )

    # 列指定を解決
    required_fields = {"item_name", "jlac10"}
    optional_fields = {"abbreviation", "jlac10_standard_name"}
    all_fields = required_fields | optional_fields

    for field in required_fields:
        if field not in column_map:
            raise ValueError(f"必須フィールド '{field}' が column_map にありません")

    resolved: dict[str, int] = {}
    for field, spec in column_map.items():
        if field not in all_fields:
            logger.warning("不明なフィールド '%s' は無視します", field)
            continue
        resolved[field] = _resolve_column_index(spec, header_row)
        logger.debug("列解決: %s = 列%d (指定: '%s')", field, resolved[field], spec)

    # データ変換
    items: list[dict] = []
    skipped = 0

    for row_num, row in enumerate(data_rows, start=skip_rows + 1):
        # 空行スキップ
        if not row or all(cell.strip() == "" for cell in row):
            logger.debug("空行スキップ: 行%d", row_num)
            skipped += 1
            continue

        # item_name と abbreviation の両方が空ならスキップ
        item_col = resolved["item_name"]
        item_name = row[item_col].strip() if item_col < len(row) else ""

        abbreviation = ""
        if "abbreviation" in resolved:
            abbr_col = resolved["abbreviation"]
            abbreviation = row[abbr_col].strip() if abbr_col < len(row) else ""

        if not item_name and not abbreviation:
            logger.debug("空行スキップ (item_name・略称ともに空): 行%d", row_num)
            skipped += 1
            continue

        # item_name が空なら略称で代用
        if not item_name and abbreviation:
            item_name = abbreviation
            logger.debug("item_name空 → 略称で代用: 行%d '%s'", row_num, abbreviation)

        # JLAC10 取得・正規化
        jlac10_col = resolved["jlac10"]
        raw_jlac10 = row[jlac10_col].strip() if jlac10_col < len(row) else ""
        jlac10 = raw_jlac10.replace("-", "")
        if raw_jlac10 != jlac10:
            logger.debug("JLAC10正規化: 行%d '%s' → '%s'", row_num, raw_jlac10, jlac10)
        jlac10_status = classify_jlac10(jlac10)

        # analyte_code: JLAC10の先頭5桁
        analyte_code = jlac10[:5] if len(jlac10) >= 5 else ""

        # abbreviation は上で既に取得済み

        jlac10_standard_name = ""
        if "jlac10_standard_name" in resolved:
            col = resolved["jlac10_standard_name"]
            jlac10_standard_name = row[col].strip() if col < len(row) else ""

        logger.debug(
            "Row %d: item_name='%s' jlac10='%s' status=%s",
            row_num, item_name, jlac10, jlac10_status,
        )

        items.append({
            "hospital": hospital,
            "item_name": item_name,
            "abbreviation": abbreviation,
            "jlac10": jlac10,
            "jlac10_status": jlac10_status,
            "analyte_code": analyte_code,
            "jlac10_standard_name": jlac10_standard_name,
        })

    # JLAC10ステータス内訳
    _status_counts: dict[str, int] = {}
    for it in items:
        s = it["jlac10_status"]
        _status_counts[s] = _status_counts.get(s, 0) + 1
    logger.debug(
        "変換サマリー: %s",
        ", ".join(f"{k}={v}" for k, v in sorted(_status_counts.items())),
    )

    logger.info(
        "変換完了: %d件 (空行スキップ: %d件, ソース: %s)",
        len(items), skipped, filepath.name,
    )

    result = {
        "metadata": {
            "hospital": hospital,
            "source_file": filepath.name,
            "exported_at": datetime.now(timezone.utc).isoformat(),
            "total_items": len(items),
        },
        "items": items,
    }

    # JSON出力
    if output_path is None:
        output_path = filepath.with_suffix(".json")
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(result, ensure_ascii=False, indent=2),
        encoding="utf-8-sig",
    )
    logger.info("JSON出力: %s", output_path)

    return result


# ---------------------------------------------------------------------------
# usage 判定（依頼 / 結果 / 依頼,結果）
# ---------------------------------------------------------------------------

def _detect_usage(sheet_name: str | None, vendor_type: str | None) -> str:
    """シート名からJLAC10の用途を判定

    Args:
        sheet_name: シート名
        vendor_type: "separate"（依頼/結果別）or "unified"（一体型）

    Returns:
        "依頼" / "結果" / "依頼,結果"
    """
    if vendor_type == "unified":
        return "依頼,結果"

    if not sheet_name:
        return "依頼,結果"

    s = sheet_name.lower()
    has_request = "依頼" in sheet_name or "request" in s or "order" in s
    has_result = "結果" in sheet_name or "result" in s

    if has_request and has_result:
        return "依頼,結果"
    if has_request:
        return "依頼"
    if has_result:
        return "結果"
    return "依頼,結果"


# ---------------------------------------------------------------------------
# ベンダー対応自動変換
# ---------------------------------------------------------------------------

def convert_auto(
    filepath: Path,
    vendor: str | None = None,
    hospital: str = "",
    sheet_name: str | None = None,
    skip_rows: int = 1,
    output_path: Path | None = None,
) -> dict:
    """ヘッダー自動検出で Excel/CSV を変換する。

    ベンダー名を指定すればそのプロファイルで列検出、
    未指定なら汎用キーワードで推定。

    Args:
        filepath: 入力ファイル
        vendor: ベンダー名 (NEC/Fujitsu/IBM/SSI/SBS/KHI/CSI/NAIS)
        hospital: 病院名
        sheet_name: シート名（省略で最初のシート）
        skip_rows: ヘッダー行数
        output_path: 出力先

    Returns:
        convert_tabular と同じ形式の dict
    """
    filepath = Path(filepath)
    suffix = filepath.suffix.lower()

    if suffix == ".xlsx":
        data_rows, header_row = _read_xlsx(filepath, sheet_name, skip_rows)
    elif suffix == ".csv":
        data_rows, header_row = _read_csv(filepath, skip_rows)
    else:
        raise ValueError(f"未対応の形式: {suffix}")

    if not header_row:
        raise ValueError("ヘッダー行が見つかりません")

    logger.info("ヘッダー自動検出: vendor=%s, sheet=%s", vendor, sheet_name)
    logger.debug("ヘッダー: %s", header_row)

    cols = detect_columns(header_row, vendor=vendor, sheet_name=sheet_name)

    logger.info("検出結果: %s",
                {k: f"{v}列目({header_row[v]})" if v is not None else "未検出"
                 for k, v in cols.items()})

    # 必須列チェック
    if cols.get("item_name") is None:
        raise ValueError(
            f"項目名列が検出できません。ヘッダー: {header_row}\n"
            f"--col-item で手動指定してください。"
        )

    # convert_tabular 用の column_map に変換
    column_map = {}
    field_mapping = {
        "item_name": "item_name",
        "jlac10": "jlac10",
        "short_name": "abbreviation",
        "jlac10_name": "jlac10_standard_name",
    }

    for detected_key, col_idx in cols.items():
        if col_idx is None:
            continue
        cm_key = field_mapping.get(detected_key)
        if cm_key:
            # 列番号をアルファベットに変換して convert_tabular に渡す
            column_map[cm_key] = str(col_idx + 1)  # 1-based 数値文字列

    # item_name は必須
    if "item_name" not in column_map:
        raise ValueError("項目名列が検出できません")

    # jlac10 がない場合はダミー（空列）
    if "jlac10" not in column_map:
        logger.warning("JLAC10列が検出されませんでした。空として処理します。")
        # 存在しない列番号を指定（全て空になる）
        column_map["jlac10"] = str(len(header_row) + 1)

    # usage 判定（依頼 / 結果 / 依頼,結果）
    from .vendor_profiles import get_vendor_info
    vendor_info = get_vendor_info(vendor) if vendor else None
    vendor_type = vendor_info.get("type") if vendor_info else None
    usage = _detect_usage(sheet_name, vendor_type)
    logger.info("用途判定: sheet='%s', vendor_type=%s → usage='%s'", sheet_name, vendor_type, usage)

    result = convert_tabular(
        filepath=filepath,
        column_map=column_map,
        hospital=hospital,
        sheet_name=sheet_name,
        skip_rows=skip_rows,
        output_path=output_path,
    )

    # usage を metadata と各 item に追加
    result["metadata"]["usage"] = usage
    result["metadata"]["source_sheet"] = sheet_name or ""
    result["metadata"]["vendor"] = vendor or ""
    for item in result["items"]:
        item["usage"] = usage
        item["source_sheet"] = sheet_name or ""

    # output_path に再書き込み（usage追加分）
    out = output_path or Path(filepath).with_suffix(".json")
    Path(out).write_text(
        json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    return result


def write_jlac10_to_excel(
    source_path: Path,
    mapping_results: list[dict],
    output_path: Path,
    sheet_name: str | None = None,
    skip_rows: int = 1,
) -> Path:
    """元の Excel に JLAC10 コードと標準名称を追記して保存

    元ファイルの最終列の後ろに2列追加:
      - JLAC10（マッピング結果）
      - JLAC10標準名称

    Args:
        source_path: 元の Excel ファイル
        mapping_results: mapper.bulk_map() の results リスト
        output_path: 出力先
        sheet_name: 対象シート名
        skip_rows: ヘッダー行数
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.load_workbook(str(source_path))
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # 最終列を取得
    max_col = ws.max_column
    jlac10_col = max_col + 1
    name_col = max_col + 2
    status_col = max_col + 3

    # ヘッダー追加
    header_font = Font(bold=True)
    ws.cell(row=skip_rows, column=jlac10_col, value="JLAC10").font = header_font
    ws.cell(row=skip_rows, column=name_col, value="JLAC10標準名称").font = header_font
    ws.cell(row=skip_rows, column=status_col, value="Status").font = header_font

    # 色定義
    fill_auto = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_candidate = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    fill_manual = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # データ書き込み
    for i, item in enumerate(mapping_results):
        row_num = skip_rows + 1 + i
        best = item.get("best_match")
        status = item.get("status", "manual")

        jlac10_val = best["jlac10"] if best else ""
        name_val = best["matched_name"] if best else ""

        ws.cell(row=row_num, column=jlac10_col, value=jlac10_val)
        ws.cell(row=row_num, column=name_col, value=name_val)
        ws.cell(row=row_num, column=status_col, value=status)

        fill = fill_auto if status == "auto" else fill_candidate if status == "candidate" else fill_manual
        for col in (jlac10_col, name_col, status_col):
            ws.cell(row=row_num, column=col).fill = fill

    # 列幅
    ws.column_dimensions[openpyxl.utils.get_column_letter(jlac10_col)].width = 20
    ws.column_dimensions[openpyxl.utils.get_column_letter(name_col)].width = 35
    ws.column_dimensions[openpyxl.utils.get_column_letter(status_col)].width = 12

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info("Excel追記完了: %s (JLAC10=%d列, 名称=%d列, Status=%d列)",
                output_path, jlac10_col, name_col, status_col)
    return output_path
