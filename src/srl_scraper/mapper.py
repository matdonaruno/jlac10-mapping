"""院内検査項目 → JLAC10 一括マッピングモジュール

入力:  converter.py が出力する items 形式のリスト
処理:  SearchIndex で各 item_name をあいまい検索しスコアで分類
       SOP / 試薬情報があれば測定法コードで候補を絞り込み
出力:  auto / candidate / manual に分類されたマッピング結果（Excel + JSON）
"""

import json
import logging
import re
from datetime import datetime, timezone
from pathlib import Path

from .search import SearchIndex, _normalize

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# SOP 情報による候補絞り込み (T9)
# ---------------------------------------------------------------------------

def enrich_with_sop(
    items: list[dict],
    sop_data: list[dict],
) -> list[dict]:
    """マッピング対象項目に SOP 情報を付加

    SOP の test_item と item の item_name をあいまいマッチし、
    一致した SOP の method_summary, reagent を項目に追加。
    """
    if not sop_data:
        return items

    # SOP をインデックス化（正規化済み名称）
    sop_index = []
    for sop in sop_data:
        name = sop.get("test_item", "")
        if name:
            sop_index.append({
                "name_norm": _normalize(name),
                "name": name,
                "method": sop.get("method_summary", ""),
                "reagent": sop.get("reagent", ""),
                "instrument": sop.get("instrument", ""),
            })

    for item in items:
        item_norm = _normalize(item.get("item_name", ""))
        if not item_norm:
            continue

        best_score = 0
        best_sop = None
        for sop in sop_index:
            # 単純な部分一致スコア
            if item_norm == sop["name_norm"]:
                score = 100
            elif item_norm in sop["name_norm"] or sop["name_norm"] in item_norm:
                score = 70
            else:
                score = 0
            if score > best_score:
                best_score = score
                best_sop = sop

        if best_sop and best_score >= 50:
            item["_sop_method"] = best_sop["method"]
            item["_sop_reagent"] = best_sop["reagent"]
            item["_sop_score"] = best_score

    return items


def adjust_scores_with_method(
    candidates: list[dict],
    sop_method: str,
    method_keyword_map: dict | None = None,
) -> list[dict]:
    """SOP 測定法情報で候補のスコアを補正

    候補の JLAC10 測定法コード(12-14桁) と SOP 測定法のキーワードが
    マッチすれば +15、不一致なら -5 でスコア補正。
    """
    if not sop_method or not method_keyword_map or not candidates:
        return candidates

    sop_norm = _normalize(sop_method)

    for cand in candidates:
        jlac10 = cand.get("jlac10", "")
        if len(jlac10) < 15:
            continue
        method_code = jlac10[12:15]
        method_info = method_keyword_map.get(method_code, {})
        keywords = method_info.get("keywords", [])

        matched = False
        for kw in keywords:
            if kw.lower() in sop_norm or _normalize(kw) in sop_norm:
                matched = True
                break

        if matched:
            cand["score"] = min(cand["score"] + 15, 100)
            cand["_method_boost"] = True
        else:
            cand["score"] = max(cand["score"] - 5, 0)

    # 再ソート
    candidates.sort(key=lambda x: -x["score"])
    return candidates


def bulk_map(
    items: list[dict],
    index: SearchIndex,
    auto_threshold: float = 90.0,
    candidate_threshold: float = 50.0,
    max_candidates: int = 5,
    sop_data: list[dict] | None = None,
    method_keyword_map: dict | None = None,
) -> dict:
    """一括マッピング実行

    各 item の item_name で index.search() を実行し、
    スコアに基づいて auto / candidate / manual に分類する。

    Args:
        items: converter.py の出力 items 形式
            各要素は {"item_name": str, ...} を含む dict
        index: search.py の SearchIndex インスタンス
        auto_threshold: この値以上なら自動マッピング (default: 90.0)
        candidate_threshold: この値以上なら候補あり (default: 50.0)
        max_candidates: 候補として返す最大件数 (default: 5)

    Returns:
        {
            "metadata": {"total": N, "auto": N, "candidate": N, "manual": N,
                         "mapped_at": "ISO8601"},
            "results": [{
                "item_name": str,
                "hospital": str,
                "abbreviation": str,
                "original_jlac10": str,
                "status": "auto" | "candidate" | "manual",
                "best_match": {...} or None,
                "candidates": [...],
            }]
        }
    """
    # SOP 情報があれば items に付加
    if sop_data:
        items = enrich_with_sop(items, sop_data)

    results: list[dict] = []
    counts = {"auto": 0, "candidate": 0, "manual": 0}

    for item in items:
        item_name = item.get("item_name", "").strip()
        if not item_name:
            results.append({
                "item_name": "",
                "hospital": item.get("hospital", ""),
                "abbreviation": item.get("abbreviation", ""),
                "original_jlac10": item.get("jlac10", ""),
                "status": "manual",
                "best_match": None,
                "candidates": [],
            })
            counts["manual"] += 1
            continue

        hits = index.search(item_name, max_results=max_candidates)

        best_match = None
        candidates = []

        if hits:
            candidates = [
                {
                    "jlac10": h["jlac10"],
                    "matched_name": h["matched_name"],
                    "score": h["score"],
                    "analyte_code": h["analyte_code"],
                }
                for h in hits
            ]

            # SOP 測定法でスコア補正
            sop_method = item.get("_sop_method", "")
            if sop_method and method_keyword_map:
                candidates = adjust_scores_with_method(candidates, sop_method, method_keyword_map)

            top = candidates[0]
            best_match = {
                "jlac10": top["jlac10"],
                "matched_name": top["matched_name"],
                "score": top["score"],
                "analyte_code": top["analyte_code"],
                "all_names": hits[0].get("all_names", []) if hits else [],
                "sources": hits[0].get("sources", {}) if hits else {},
            }

        if best_match and best_match["score"] >= auto_threshold:
            status = "auto"
        elif best_match and best_match["score"] >= candidate_threshold:
            status = "candidate"
        else:
            status = "manual"

        counts[status] += 1

        results.append({
            "item_name": item_name,
            "hospital": item.get("hospital", ""),
            "abbreviation": item.get("abbreviation", ""),
            "original_jlac10": item.get("jlac10", ""),
            "status": status,
            "best_match": best_match,
            "candidates": candidates,
        })

    total = len(results)
    logger.info(
        "マッピング完了: %d件 (auto=%d, candidate=%d, manual=%d)",
        total, counts["auto"], counts["candidate"], counts["manual"],
    )

    return {
        "metadata": {
            "total": total,
            "auto": counts["auto"],
            "candidate": counts["candidate"],
            "manual": counts["manual"],
            "auto_threshold": auto_threshold,
            "candidate_threshold": candidate_threshold,
            "mapped_at": datetime.now(timezone.utc).isoformat(),
        },
        "results": results,
    }


def export_mapping_excel(results: dict, output_path: Path) -> None:
    """マッピング結果を色付き Excel に出力する。

    列構成:
        Status | Item Name | Original JLAC10 | Matched JLAC10 |
        Matched Name | Score | Analyte Code | Alt1 JLAC10 | Alt2 JLAC10

    色分け:
        auto     = 緑背景 (C6EFCE)
        candidate = 黄背景 (FFEB9C)
        manual   = 赤背景 (FFC7CE)
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mapping Results"

    # ヘッダー定義
    headers = [
        "Status",
        "Item Name",
        "Abbreviation",
        "Original JLAC10",
        "Matched JLAC10",
        "Matched Name",
        "Score",
        "Analyte Code",
        "Alt1 JLAC10",
        "Alt2 JLAC10",
    ]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 行の色分け
    fill_auto = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_candidate = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    fill_manual = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    status_fill_map = {
        "auto": fill_auto,
        "candidate": fill_candidate,
        "manual": fill_manual,
    }

    for row_idx, entry in enumerate(results["results"], start=2):
        status = entry["status"]
        best = entry.get("best_match")
        candidates = entry.get("candidates", [])

        matched_jlac10 = best["jlac10"] if best else ""
        matched_name = best["matched_name"] if best else ""
        score = best["score"] if best else ""
        analyte_code = best["analyte_code"] if best else ""

        # 代替候補（best_match 以外の上位2件）
        alts = [c for c in candidates if c["jlac10"] != matched_jlac10]
        alt1_jlac10 = alts[0]["jlac10"] if len(alts) > 0 else ""
        alt2_jlac10 = alts[1]["jlac10"] if len(alts) > 1 else ""

        row_data = [
            status,
            entry["item_name"],
            entry.get("abbreviation", ""),
            entry.get("original_jlac10", ""),
            matched_jlac10,
            matched_name,
            score,
            analyte_code,
            alt1_jlac10,
            alt2_jlac10,
        ]

        fill = status_fill_map.get(status, fill_manual)
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill

    # 列幅の自動調整
    column_widths = [10, 30, 15, 18, 18, 30, 8, 14, 18, 18]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # オートフィルタ
    ws.auto_filter.ref = ws.dimensions

    # メタデータシート
    ws_meta = wb.create_sheet("Metadata")
    meta = results["metadata"]
    meta_rows = [
        ("Total Items", meta["total"]),
        ("Auto Mapped", meta["auto"]),
        ("Candidates", meta["candidate"]),
        ("Manual Required", meta["manual"]),
        ("Auto Threshold", meta["auto_threshold"]),
        ("Candidate Threshold", meta["candidate_threshold"]),
        ("Mapped At", meta["mapped_at"]),
    ]
    for row_idx, (key, val) in enumerate(meta_rows, start=1):
        ws_meta.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
        ws_meta.cell(row=row_idx, column=2, value=val)
    ws_meta.column_dimensions["A"].width = 20
    ws_meta.column_dimensions["B"].width = 30

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info("Excel出力: %s", output_path)


def export_mapping_json(results: dict, output_path: Path) -> None:
    """マッピング結果を JSON に出力する。"""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(results, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    logger.info("JSON出力: %s", output_path)
