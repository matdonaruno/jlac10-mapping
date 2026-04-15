"""JSLM JLAC10コード表 取得・パーサー

毎月更新される JLAC10 コード表 xlsx を取得し、
分析物・識別・材料・測定法・結果識別の各コードマスターを JSON 化する。

ソース: https://www.jslm.org/committees/code/index.html
"""

import json
import logging
import re
import time
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import requests
from bs4 import BeautifulSoup

from .scraper import HEADERS, REQUEST_INTERVAL, REQUEST_TIMEOUT

logger = logging.getLogger(__name__)

JSLM_URL = "https://www.jslm.org/committees/code/index.html"
JSLM_BASE = "https://www.jslm.org/committees/code/"
JSLM_LAST_UPDATE_FILE = "jslm_last_update.txt"


def discover_latest_xlsx(session: requests.Session) -> tuple[str, str]:
    """JSLM ページから最新の JLAC10 臨床検査コード表 xlsx URL と版数を取得"""
    resp = session.get(JSLM_URL, headers=HEADERS, timeout=REQUEST_TIMEOUT)
    resp.raise_for_status()
    resp.encoding = resp.apparent_encoding or "utf-8"
    soup = BeautifulSoup(resp.text, "lxml")

    for tr in soup.find_all("tr"):
        th = tr.find("th")
        if not th or "JLAC10コード表_臨床検査" not in th.get_text():
            continue
        cells = tr.find_all("td")
        if len(cells) < 3:
            continue
        update_date = cells[0].get_text(strip=True)  # "2026/03"
        version = cells[1].get_text(strip=True)  # "136"
        link = tr.find("a", href=re.compile(r"\.xlsx$"))
        if link:
            href = link.get("href")
            url = f"{JSLM_BASE}{href}" if not href.startswith("http") else href
            return url, f"{version}_{update_date}"

    raise RuntimeError("JLAC10コード表の xlsx リンクが見つかりません")


def check_jslm_update_needed(output_dir: Path, session: requests.Session) -> tuple[bool, str, str]:
    """JSLM の更新有無を確認"""
    _, remote_version = discover_latest_xlsx(session)
    date_file = output_dir / JSLM_LAST_UPDATE_FILE
    local_version = ""
    if date_file.exists():
        local_version = date_file.read_text(encoding="utf-8").strip()
    return remote_version != local_version, remote_version, local_version


def download_xlsx(url: str, session: requests.Session, output_dir: Path) -> Path:
    """xlsx をダウンロード"""
    resp = session.get(url, headers=HEADERS, timeout=60)
    resp.raise_for_status()
    filename = url.split("/")[-1]
    filepath = output_dir / filename
    filepath.write_bytes(resp.content)
    logger.info("JLAC10コード表ダウンロード: %s (%d bytes)", filepath, len(resp.content))
    return filepath


def parse_analyte_codes(ws) -> list[dict]:
    """分析物コードシートをパース"""
    items = []
    current_category = ""
    current_category_name = ""
    for row in ws.iter_rows(min_row=6, values_only=True):
        code = str(row[2]).strip() if row[2] else ""
        if not code or len(code) != 5:
            continue
        if row[0]:
            current_category = str(row[0]).strip()
        if row[1]:
            current_category_name = str(row[1]).strip()
        items.append({
            "code": code,
            "name": str(row[3] or "").strip(),
            "name2": str(row[4] or "").strip(),
            "name_en": str(row[5] or "").strip(),
            "category": current_category,
            "category_name": current_category_name,
        })
    return items


def parse_identification_codes(ws) -> list[dict]:
    """識別コードシートをパース"""
    items = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        code = str(row[1]).strip() if row[1] else ""
        if not code or not re.match(r"^\d{4}$", code):
            continue
        items.append({
            "code": code,
            "name": str(row[2] or "").strip(),
            "name2": str(row[3] or "").strip(),
            "name_en": str(row[4] or "").strip(),
        })
    return items


def parse_material_codes(ws) -> list[dict]:
    """材料コードシートをパース"""
    items = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        code = str(row[5]).strip() if row[5] else ""
        if not code or not re.match(r"^\d{3}$", code):
            continue
        items.append({
            "code": code,
            "name": str(row[6] or "").strip(),
            "name_en": str(row[7] or "").strip(),
        })
    return items


def parse_method_codes(ws) -> list[dict]:
    """測定法コードシートをパース"""
    items = []
    current_cat1 = ""
    current_cat2 = ""
    for row in ws.iter_rows(min_row=6, values_only=True):
        code = str(row[2]).strip() if row[2] else ""
        if not code or not re.match(r"^\d{3}$", code):
            continue
        if row[0]:
            current_cat1 = str(row[0]).strip()
        if row[1]:
            current_cat2 = str(row[1]).strip()
        items.append({
            "code": code,
            "name": str(row[3] or "").strip(),
            "name2": str(row[4] or "").strip(),
            "name_en": str(row[5] or "").strip(),
            "category": current_cat1,
            "subcategory": current_cat2,
        })
    return items


def parse_result_common_codes(ws) -> list[dict]:
    """結果識別コード（共通）シートをパース"""
    items = []
    current_category = ""
    for row in ws.iter_rows(min_row=6, values_only=True):
        code = str(row[1]).strip() if row[1] else ""
        if not code or not re.match(r"^\d{2,3}$", code):
            continue
        if row[0]:
            current_category = str(row[0]).strip()
        items.append({
            "code": code.zfill(3),
            "name": str(row[2] or "").strip(),
            "category": current_category,
        })
    return items


def parse_result_specific_codes(ws) -> list[dict]:
    """結果識別コード（固有）シートをパース"""
    items = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        analyte = str(row[1]).strip() if row[1] else ""
        ident = str(row[2]).strip() if row[2] else ""
        code = str(row[3]).strip() if row[3] else ""
        if not analyte or not code:
            continue
        items.append({
            "analyte_code": analyte,
            "identification_code": ident,
            "code": code.zfill(2),
            "name": str(row[4] or "").strip(),
            "name_en": str(row[5] or "").strip(),
        })
    return items


def parse_xlsx(filepath: Path) -> dict:
    """JLAC10コード表 xlsx をパースして全コードマスターを返す"""
    wb = openpyxl.load_workbook(filepath, read_only=True)

    result = {
        "analyte": parse_analyte_codes(wb["分析物コード"]),
        "identification": parse_identification_codes(wb["識別コード "]),
        "material": parse_material_codes(wb["材料コード"]),
        "method": parse_method_codes(wb["測定法コード"]),
        "result_common": parse_result_common_codes(wb["結果識別コード表（共通コード）"]),
        "result_specific": parse_result_specific_codes(wb["結果識別（固有コード）"]),
    }

    wb.close()
    return result


def build_lookup(master: dict) -> dict:
    """コードマスターから高速検索用の辞書を構築"""
    return {
        "analyte": {item["code"]: item for item in master["analyte"]},
        "identification": {item["code"]: item for item in master["identification"]},
        "material": {item["code"]: item for item in master["material"]},
        "method": {item["code"]: item for item in master["method"]},
        "result_common": {item["code"]: item for item in master["result_common"]},
    }


def decode_jlac10(code: str, lookup: dict) -> dict:
    """JLAC10コード(15桁ハイフンなし)を分解してコード名称を付与"""
    code = code.replace("-", "")
    if len(code) < 15:
        return {"raw": code, "valid": False}

    analyte = code[0:5]
    identification = code[5:9]
    material = code[9:12]
    method = code[12:15]

    a = lookup.get("analyte", {}).get(analyte, {})
    i = lookup.get("identification", {}).get(identification, {})
    m = lookup.get("material", {}).get(material, {})
    mt = lookup.get("method", {}).get(method, {})

    return {
        "raw": code,
        "valid": True,
        "analyte": {
            "code": analyte,
            "name": a.get("name", ""),
            "name_en": a.get("name_en", ""),
        },
        "identification": {
            "code": identification,
            "name": i.get("name", ""),
        },
        "material": {
            "code": material,
            "name": m.get("name", ""),
        },
        "method": {
            "code": method,
            "name": mt.get("name", ""),
        },
    }


def scrape_all(
    output_dir: Path | None = None,
    check_update: bool = False,
) -> dict | None:
    """JSLM JLAC10コード表を取得してJSON化"""
    output_dir = output_dir or Path("data")
    output_dir.mkdir(parents=True, exist_ok=True)

    session = requests.Session()
    now = datetime.now(timezone.utc)

    xlsx_url, version = discover_latest_xlsx(session)
    logger.info("JSLM: 最新コード表 version=%s, URL=%s", version, xlsx_url)

    if check_update:
        date_file = output_dir / JSLM_LAST_UPDATE_FILE
        local_version = date_file.read_text(encoding="utf-8").strip() if date_file.exists() else ""
        if version == local_version:
            logger.info("JSLM: 更新なし (version: %s)。スキップします。", version)
            return None
        logger.info("JSLM: 更新を検出 (%s → %s)", local_version or "(初回)", version)

    time.sleep(REQUEST_INTERVAL)

    # ダウンロード
    xlsx_path = download_xlsx(xlsx_url, session, output_dir)

    # パース
    master = parse_xlsx(xlsx_path)

    # 検索用辞書
    lookup = build_lookup(master)

    result = {
        "metadata": {
            "source": JSLM_URL,
            "xlsx_url": xlsx_url,
            "version": version,
            "scraped_at": now.isoformat(),
            "counts": {k: len(v) for k, v in master.items()},
        },
        "master": master,
    }

    # JSON保存
    filepath = output_dir / "jlac10_master.json"
    filepath.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    logger.info("JSLM 保存完了: %s", filepath)

    # 検索用辞書も保存
    lookup_path = output_dir / "jlac10_lookup.json"
    lookup_path.write_text(json.dumps(lookup, ensure_ascii=False, indent=2), encoding="utf-8")

    # バージョン記録
    (output_dir / JSLM_LAST_UPDATE_FILE).write_text(version, encoding="utf-8")

    return result
