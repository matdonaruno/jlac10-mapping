"""設定依頼フォーマット出力

SSMIXパース結果やマッピング結果から、ベンダー別の設定依頼Excelを生成する。
"""

import logging
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .jslm import decode_jlac10

logger = logging.getLogger(__name__)

# ヘッダースタイル
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

# データセルスタイル
DATA_FONT = Font(size=10)
DATA_ALIGNMENT = Alignment(vertical="center", wrap_text=False)


def _apply_header_style(ws, headers: list[str]) -> None:
    """ヘッダー行にスタイルを適用"""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT


def _auto_column_width(ws) -> None:
    """列幅を自動調整"""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                # 日本語は2文字分としてカウント
                val = str(cell.value)
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                max_length = max(max_length, length)
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width


def generate_jlac10_standard_name(jlac10: str, lookup: dict) -> str:
    """JLAC10コードから標準名称を生成

    形式: {分析物名}_{材料名}_{測定法名}_{結果識別名}
    例: 膵アミラーゼ_血清_可視吸光光度法_定量値

    jslm.py の decode_jlac10() を使ってパート名を取得し連結。
    """
    if not jlac10 or not lookup:
        return ""

    decoded = decode_jlac10(jlac10, lookup)
    if not decoded.get("valid"):
        return ""

    parts = []
    analyte_name = decoded.get("analyte", {}).get("name", "")
    material_name = decoded.get("material", {}).get("name", "")
    method_name = decoded.get("method", {}).get("name", "")
    identification_name = decoded.get("identification", {}).get("name", "")

    if analyte_name:
        parts.append(analyte_name)
    if material_name:
        parts.append(material_name)
    if method_name:
        parts.append(method_name)
    if identification_name:
        parts.append(identification_name)

    return "_".join(parts)


def export_request_format(items: list[dict], output_path: Path,
                          vendor: str = "", usage: str = "依頼") -> Path:
    """依頼用フォーマットExcel出力

    列: ローカルコード | 項目名称 | CS名 | JLAC10 | JLAC10標準名称

    items: [{local_code, item_name, cs_name, jlac10, jlac10_standard_name}, ...]
    usage: "依頼" or "結果"
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{usage}フォーマット"

    headers = ["ローカルコード", "項目名称", "CS名", "JLAC10", "JLAC10標準名称"]
    _apply_header_style(ws, headers)

    for row_idx, item in enumerate(items, 2):
        ws.cell(row=row_idx, column=1, value=item.get("local_code", ""))
        ws.cell(row=row_idx, column=2, value=item.get("item_name", ""))
        ws.cell(row=row_idx, column=3, value=item.get("cs_name", ""))
        ws.cell(row=row_idx, column=4, value=item.get("jlac10", ""))
        ws.cell(row=row_idx, column=5, value=item.get("jlac10_standard_name", ""))

        for col_idx in range(1, 6):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGNMENT

    _auto_column_width(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info("%s フォーマット出力: %s (%d件)", usage, output_path, len(items))
    return output_path


def export_result_format(items: list[dict], output_path: Path) -> Path:
    """結果用フォーマットExcel出力（依頼用と同じ列構成）"""
    return export_request_format(items, output_path, usage="結果")


def _truncate_bytes(text: str, max_bytes: int) -> str:
    """UTF-8でmax_bytesを超えないよう切り詰め"""
    encoded = text.encode("utf-8")
    if len(encoded) <= max_bytes:
        return text
    # 1文字ずつ減らして確認
    while len(text.encode("utf-8")) > max_bytes:
        text = text[:-1]
    return text


def _build_hl7key(item: dict, exam_type: str) -> str:
    """exam_typeに応じてHL7KEYを生成

    ルール:
    - 通常(検体): ローカルコード
    - 細菌結果: "OTKKINF" + コード
    - 塗抹: "ER02TMTCD" + 菌コード
    - 同定: "ER02DTKCD" + 菌コード
    - 抗酸菌塗抹: "ER03TMTCD"
    - 抗酸菌同定: "ER03DTKCD"
    - 抗菌薬: "KJYKCD" + 菌コード
    """
    local_code = item.get("local_code", "")
    bacteria_code = item.get("bacteria_code", local_code)

    mapping = {
        "検体": local_code,
        "細菌結果": f"OTKKINF{bacteria_code}",
        "塗抹": f"ER02TMTCD{bacteria_code}",
        "同定": f"ER02DTKCD{bacteria_code}",
        "抗酸菌塗抹": "ER03TMTCD",
        "抗酸菌同定": "ER03DTKCD",
        "抗菌薬": f"KJYKCD{bacteria_code}",
    }
    return mapping.get(exam_type, local_code)


def export_jj_format(items: list[dict], output_path: Path,
                     exam_type: str = "検体") -> Path:
    """富士通 JJマスタ形式Excel出力

    列: HL7KEY | ICODE | JJCODE | JJNAME

    ルール:
    - ICODE: データなしなら 'L'
    - JJNAME: 128Byte制限（超える場合は切り詰め）
    - HL7KEY: exam_typeに応じて生成
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "JJマスタ"

    headers = ["HL7KEY", "ICODE", "JJCODE", "JJNAME"]
    _apply_header_style(ws, headers)

    for row_idx, item in enumerate(items, 2):
        hl7key = _build_hl7key(item, exam_type)
        icode = item.get("icode", "") or "L"
        jjcode = item.get("jlac10", "")
        jjname = item.get("item_name", "")
        # JJNAME: 128Byte制限
        jjname = _truncate_bytes(jjname, 128)

        ws.cell(row=row_idx, column=1, value=hl7key)
        ws.cell(row=row_idx, column=2, value=icode)
        ws.cell(row=row_idx, column=3, value=jjcode)
        ws.cell(row=row_idx, column=4, value=jjname)

        for col_idx in range(1, 5):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGNMENT

    _auto_column_width(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info("JJマスタ出力: %s (%d件)", output_path, len(items))
    return output_path


def export_delivery(
    items: list[dict],
    output_dir: Path,
    vendor: str,
    lookup: dict,
    hospital: str = "",
    issue_number: str = "",
) -> dict:
    """ベンダーに応じた全フォーマットを一括出力

    itemsの各要素: {
        local_code: str,     # ローカルコード
        item_name: str,      # 院内項目名称
        cs_name: str,        # CS名(99Z14等)
        jlac10: str,         # マッピング結果のJLAC10(17桁)
        usage: str,          # "依頼" or "結果"
        exam_type: str,      # "検体" or "細菌結果" etc.
    }

    出力:
      {hospital}_{issue}_依頼.xlsx
      {hospital}_{issue}_結果.xlsx
      {hospital}_{issue}_JJ.xlsx（富士通のみ）

    Returns: {"files": [出力ファイルパスリスト], "summary": {...}}
    """
    output_dir.mkdir(parents=True, exist_ok=True)

    prefix = ""
    if hospital:
        prefix += hospital
    if issue_number:
        prefix += f"_{issue_number}" if prefix else issue_number
    if not prefix:
        prefix = "delivery"

    # JLAC10標準名称を生成
    for item in items:
        if not item.get("jlac10_standard_name"):
            item["jlac10_standard_name"] = generate_jlac10_standard_name(
                item.get("jlac10", ""), lookup
            )

    # 依頼/結果に分割
    request_items = [it for it in items if it.get("usage", "依頼") == "依頼"]
    result_items = [it for it in items if it.get("usage") == "結果"]

    files: list[Path] = []

    # 依頼フォーマット出力
    if request_items:
        req_path = output_dir / f"{prefix}_依頼.xlsx"
        export_request_format(request_items, req_path, vendor=vendor, usage="依頼")
        files.append(req_path)

    # 結果フォーマット出力
    if result_items:
        res_path = output_dir / f"{prefix}_結果.xlsx"
        export_result_format(result_items, res_path)
        files.append(res_path)

    # 富士通のみJJマスタ出力
    vendor_lower = vendor.lower() if vendor else ""
    if "fujitsu" in vendor_lower or "富士通" in vendor_lower:
        jj_path = output_dir / f"{prefix}_JJ.xlsx"
        # exam_typeごとにグルーピングして出力
        exam_type = items[0].get("exam_type", "検体") if items else "検体"
        export_jj_format(items, jj_path, exam_type=exam_type)
        files.append(jj_path)

    summary = {
        "vendor": vendor,
        "hospital": hospital,
        "issue_number": issue_number,
        "total_items": len(items),
        "request_items": len(request_items),
        "result_items": len(result_items),
        "jj_exported": "fujitsu" in vendor_lower or "富士通" in vendor_lower,
        "files_count": len(files),
    }

    logger.info(
        "設定依頼出力完了: vendor=%s, hospital=%s, issue=%s, files=%d",
        vendor, hospital, issue_number, len(files),
    )

    return {"files": [str(f) for f in files], "summary": summary}
