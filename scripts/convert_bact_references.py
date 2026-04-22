"""細菌検査マッピング参照資料を Excel → JSON に変換して data/ 配下に配置。

変換対象:
  - NCDA/JANIS菌名縦積み.xlsx              → data/janis_species.json
  - NCDA/JANIS抗菌薬縦積み.xlsx            → data/janis_antibiotics.json
  - NCDA/JLAC10_検体細菌材料Lookup.xlsx    → data/bact_materials.json
  - NCDA/Multi-parameter gene-related testing.xlsx → data/gene_panels.json

方針:
  - NCDA/ は .gitignore 対象（原本 Excel は GHE に乗せない）
  - data/*.json は Git 管理し、Pages/CLI の共通参照とする
  - 縦積みの自動成長は `append_*` 関数を経由して JSON に追記する
"""

from __future__ import annotations

import json
import sys
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
NCDA = ROOT / "NCDA"
DATA = ROOT / "data"
TODAY = date.today().isoformat()


def _s(v: Any) -> str:
    """セル値を素直に str 化。None/空白は空文字列。"""
    if v is None:
        return ""
    return str(v).strip()


def _pad(v: Any, width: int) -> str:
    """数値/文字列のコードを 0 埋めで width 桁に整える。空は空文字列のまま。"""
    s = _s(v)
    if not s:
        return ""
    # "1011.0" のような浮動小数点文字列化も想定して整数化
    try:
        n = int(float(s))
        return str(n).zfill(width)
    except (ValueError, TypeError):
        return s


def _write_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
        f.write("\n")


def convert_janis_species() -> int:
    src = NCDA / "JANIS菌名縦積み.xlsx"
    wb = load_workbook(src, read_only=True, data_only=True)
    ws = wb["菌名"]
    entries: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        inhouse = _s(row[0])
        janis_name = _s(row[1])
        janis_code = _pad(row[2], 4)
        note = _s(row[3])
        if not (inhouse or janis_name or janis_code):
            continue
        entries.append({
            "inhouse": inhouse,
            "janis_name": janis_name,
            "janis_code": janis_code,
            "note": note,
        })
    wb.close()
    payload = {
        "version": TODAY,
        "source": "NCDA/JANIS菌名縦積み.xlsx",
        "schema": {
            "inhouse": "院内表記（表記ゆれ含む、修正せずそのまま積む）",
            "janis_name": "JANIS菌名（基準表記）",
            "janis_code": "JANIS菌名コード(4桁、0埋め)",
            "note": "備考",
        },
        "rules": {
            "9999": "菌名でない項目（コメントのみ）",
            "9998": "その他の菌種（同定不能含む）",
        },
        "count": len(entries),
        "entries": entries,
    }
    _write_json(DATA / "janis_species.json", payload)
    return len(entries)


def convert_janis_antibiotics() -> int:
    src = NCDA / "JANIS抗菌薬縦積み.xlsx"
    wb = load_workbook(src, read_only=True, data_only=True)
    ws = wb["抗菌薬_縦積み"]
    entries: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        inhouse = _s(row[0])
        janis_abbr = _s(row[1])
        janis_name = _s(row[2])
        janis_code = _pad(row[3], 4)
        note = _s(row[4])
        if not (inhouse or janis_abbr or janis_name or janis_code):
            continue
        entries.append({
            "inhouse": inhouse,
            "janis_abbr": janis_abbr,
            "janis_name": janis_name,
            "janis_code": janis_code,
            "note": note,
        })
    wb.close()
    payload = {
        "version": TODAY,
        "source": "NCDA/JANIS抗菌薬縦積み.xlsx",
        "schema": {
            "inhouse": "院内表記（濃度・マーカー混入含む、修正せずそのまま積む）",
            "janis_abbr": "JANIS抗菌薬略号",
            "janis_name": "JANIS抗菌薬名（日本語、略号付き）",
            "janis_code": "JANIS抗菌薬コード(4桁、0埋め)",
            "note": "備考（検索注意項目等、略号衝突警告を含む）",
        },
        "count": len(entries),
        "entries": entries,
    }
    _write_json(DATA / "janis_antibiotics.json", payload)
    return len(entries)


def convert_bact_materials() -> int:
    src = NCDA / "JLAC10_検体細菌材料Lookup.xlsx"
    wb = load_workbook(src, read_only=True, data_only=True)
    ws = wb["細菌材料"]
    entries: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        material_name = _s(row[0])
        jlac10_material = _pad(row[1], 3)
        jlac10_standard_name = _s(row[2])
        if not (material_name or jlac10_material or jlac10_standard_name):
            continue
        entries.append({
            "material_name": material_name,
            "jlac10_material": jlac10_material,
            "jlac10_standard_name": jlac10_standard_name,
        })
    wb.close()
    payload = {
        "version": TODAY,
        "source": "NCDA/JLAC10_検体細菌材料Lookup.xlsx",
        "schema": {
            "material_name": "病院側の材料名",
            "jlac10_material": "JLAC10材料コード(3桁)。運用上組合せ可能なら 'xxx'",
            "jlac10_standard_name": "JLAC10材料コードの標準名称",
        },
        "count": len(entries),
        "entries": entries,
    }
    _write_json(DATA / "bact_materials.json", payload)
    return len(entries)


def convert_gene_panels() -> int:
    src = NCDA / "Multi-parameter gene-related testing.xlsx"
    wb = load_workbook(src, read_only=True, data_only=True)
    sheets: dict[str, list[dict]] = {}

    # コード表: 商品名/一般名/分類コード/企業名/分析物/識別/材料/測定法/結果
    ws = wb["コード表"]
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    panels = []
    for row in rows:
        if all(v is None or _s(v) == "" for v in row):
            continue
        entry = {
            "product_name": _s(row[1]),
            "generic_name": _s(row[2]),
            "classification_code": _s(row[3]),
            "company": _s(row[4]),
            "analyte": _s(row[5]),
            "identification": _pad(row[6], 4),
            "material": _pad(row[7], 3),
            "methodology": _pad(row[8], 3),
            "result_identifier": _pad(row[9], 2),
        }
        if any(entry.values()):
            panels.append(entry)
    sheets["code_table"] = panels

    # コード表２: 商品名/備考/JLAC10コード/JLAC10材料/JLAC10測定方法
    ws = wb["コード表２"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    panels2 = []
    for row in rows:
        if all(v is None or _s(v) == "" for v in row):
            continue
        entry = {
            "product_name": _s(row[0]),
            "note": _s(row[1]),
            "jlac10": _s(row[2]),
            "jlac10_material_name": _s(row[3]),
            "jlac10_methodology_name": _s(row[4]),
        }
        if any(entry.values()):
            panels2.append(entry)
    sheets["code_table_2"] = panels2

    # map: 結果/項目名称/JLAC10/JLAC10標準名称（複数ブロックあり）
    ws = wb["map"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    maps = []
    for row in rows:
        if all(v is None or _s(v) == "" for v in row):
            continue
        entry = {
            "result_code": _s(row[0]),
            "item_name": _s(row[1]),
            "jlac10": _s(row[2]),
            "jlac10_standard_name": _s(row[3]),
            "cs": _s(row[4]) if len(row) > 4 else "",
        }
        # ヘッダ行（"結果"や"項目名称"だけのもの）はスキップ
        if entry["item_name"] in ("項目名称", "") and entry["jlac10"] in ("JLAC10", ""):
            continue
        if any(entry.values()):
            maps.append(entry)
    sheets["map"] = maps

    wb.close()
    payload = {
        "version": TODAY,
        "source": "NCDA/Multi-parameter gene-related testing.xlsx",
        "schema": {
            "code_table": "メーカーキット毎のJLAC10要素分解（分析物/識別/材料/測定法/結果識別）",
            "code_table_2": "商品名→JLAC10(17桁)連結形式",
            "map": "院内コード/項目名称→JLAC10 のマッピング実例",
        },
        "counts": {name: len(rows) for name, rows in sheets.items()},
        "sheets": sheets,
    }
    _write_json(DATA / "gene_panels.json", payload)
    return sum(len(v) for v in sheets.values())


def main() -> int:
    DATA.mkdir(parents=True, exist_ok=True)
    missing = [f for f in [
        NCDA / "JANIS菌名縦積み.xlsx",
        NCDA / "JANIS抗菌薬縦積み.xlsx",
        NCDA / "JLAC10_検体細菌材料Lookup.xlsx",
        NCDA / "Multi-parameter gene-related testing.xlsx",
    ] if not f.exists()]
    if missing:
        print(f"参照資料が見つかりません: {missing}", file=sys.stderr)
        return 1

    print(f"[convert] NCDA/ → data/ (version={TODAY})")
    n1 = convert_janis_species()
    print(f"  janis_species.json: {n1:,} entries")
    n2 = convert_janis_antibiotics()
    print(f"  janis_antibiotics.json: {n2:,} entries")
    n3 = convert_bact_materials()
    print(f"  bact_materials.json: {n3:,} entries")
    n4 = convert_gene_panels()
    print(f"  gene_panels.json: {n4:,} entries (3 sheets)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
